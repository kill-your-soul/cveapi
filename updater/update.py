import hashlib
import http
import json
import re
import time
from io import BytesIO
from shutil import rmtree
from typing import Any
from zipfile import ZipFile

import requests  # type: ignore  # noqa: PGH003
import untangle
from git import Repo, Tree
from openpyxl import load_workbook

from config import celery, settings
from models import Bdu, Cve, Cwe, Nvd
from utils import extract_links_from_file


@celery.task
def update_cwe() -> None:
    print("Importing CWE list...")

    # Download the file
    # with timed_operation("Downloading {}...".format(settings.MITRE_CWE_URL)):
    resp = requests.get(settings.MITRE_CWE_URL).content

    # Parse weaknesses
    # with timed_operation("Parsing cwes..."):
    z = ZipFile(BytesIO(resp))
    raw = z.open(z.namelist()[0]).read()
    obj = untangle.parse(raw.decode("utf-8"))
    weaknesses = obj.Weakness_Catalog.Weaknesses.Weakness
    categories = obj.Weakness_Catalog.Categories.Category

    # Create the objects
    cwes = {}
    # with timed_operation("Creating mappings..."):
    count = 0
    cwes = []
    for c in weaknesses + categories:
        tmp = {
            # "id": get_uuid(),
            "cwe_id": f"CWE-{c['ID']}",
            "name": c["Name"],
            "description": c.Description.cdata if hasattr(c, "Description") else c.Summary.cdata,
        }
        cwes.append(tmp)
        # resp = requests.post(settings.CVE_API_URL + "api/v1/cwe/", data=json.dumps(tmp))
        # count += 1
        # print(cwes[c["ID"]])
        # print(tmp)
    session = requests.Session()
    for item in cwes:
        cwe = Cwe(**item)
        nvd_in_hash_sum = hashlib.sha256(
            json.dumps(cwe.model_dump(), sort_keys=True).encode("utf-8"),
        ).hexdigest()
        server_data: requests.Response = session.get(settings.CVE_API_URL + f"api/v1/cwe/cwe_id/{item['cwe_id']}")
        # print(server_data["bdu"]["hash_sum"])
        if server_data.status_code == 404:
            resp = requests.post(settings.CVE_API_URL + "api/v1/cwe/", data=json.dumps(cwe))
            continue
        if server_data.json()["hash_sum"] != nvd_in_hash_sum:
            print("New data")
            resp = session.put(
                settings.CVE_API_URL + f'api/v1/cwe/{server_data["id"]}',
                data=json.dumps(item),
            )
            print(resp.text)
        else:
            print("Old data")
    # Insert the objects in database
    # with timed_operation("Inserting CWE..."):
    # db.session.bulk_insert_mappings(Cwe, cwes.values())
    # db.session.commit()
    # for cwe in cwes:
    #     print(cwe)
    # print(f"{count} CWE imported.")


@celery.task
def update_bdu() -> None:
    print("Getting info from BDU")
    # r = requests.get(settings.CVE_API_URL)
    # print(r)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    }
    xlsx = requests.get(settings.BDU_XLSX_URL, verify=False, allow_redirects=True, headers=headers).content  # noqa: S501
    wb = load_workbook(filename=BytesIO(xlsx))
    results = []
    sheet = wb.active
    session = requests.Session()
    # f = open('res.txt', 'w')
    for row in sheet.iter_rows(min_row=4):
        bdu_id = row[0].value
        description = row[1].value
        other_id = row[18].value
        # Создаем новый словарь для каждой итерации
        bdu_data = {"cve_id": "", "bdu_id": bdu_id, "description": description}
        if other_id:
            cve_match = re.findall(r"CVE-\d{4}-\d{4,7}", str(other_id))
            for cve_id in cve_match:
                # Создаем новый словарь для каждого cve_id
                temp_bdu_data = bdu_data.copy()
                temp_bdu_data["cve_id"] = cve_id
                results.append(temp_bdu_data)
        else:
            results.append(bdu_data)
    for item in results:
        bdu = Bdu(**item)
        nvd_in_hash_sum = hashlib.sha256(
            json.dumps(bdu.model_dump(), sort_keys=True).encode("utf-8"),
        ).hexdigest()
        server_data = session.get(settings.CVE_API_URL + f"api/v1/cve/?cve_id={item['cve_id']}").json()
        # print(server_data["bdu"]["hash_sum"])
        if not server_data["bdu"]:
            _resp = session.post(settings.CVE_API_URL + "api/v1/bdu/", data=json.dumps(bdu))
            continue
        if server_data["bdu"]["hash_sum"] != nvd_in_hash_sum:
            print("New data")
            resp = session.put(
                settings.CVE_API_URL + f'api/v1/bdu/{server_data["bdu"]["id"]}',
                data=json.dumps(item),
            )
            print(resp.text)
        else:
            print("Old data")
        # _resp = session.post(settings.CVE_API_URL + "api/v1/bdu/", data=json.dumps(bdu))
        # print(resp)


@celery.task
def update_nvd() -> None:  # noqa: PLR0915, PLR0912, C901
    """Import the CVE list.

    Important notice:
        This product uses data from the NVD API but is not endorsed or certified by the NVD.
    """
    print("Gettins info from NVD")
    mappings = {"vendors": {}, "products": {}, "cves": [], "changes": []}
    url_template = settings.NVD_API_URL + "?startIndex={idx}"

    start_index = 0
    total_results = 0

    while start_index <= total_results:
        url = url_template.format(idx=start_index)
        print(f"Downloading {url}")
        resp = requests.get(url)

        if not resp.ok:
            print(f"Bad response: {resp.status_code}, sleeping before retrying")
            time.sleep(10)
            continue
        data = resp.json()
        total_results = data.get("totalResults")
        for vulnerability in data.get("vulnerabilities"):
            cve_data = vulnerability.get("cve")
            cve_id = cve_data["id"]
            if "cvssMetricV31" in cve_data["metrics"]:
                cvss3 = cve_data.get("metrics")["cvssMetricV31"][0]["cvssData"]["baseScore"]
            elif "cvssMetricV30" in cve_data["metrics"]:
                cvss3 = cve_data.get("metrics")["cvssMetricV30"][0]["cvssData"]["baseScore"]
            else:
                cvss3 = 0

            if "cvssMetricV2" in cve_data.get("metrics"):
                cvss2 = cve_data.get("metrics")["cvssMetricV2"][0]["cvssData"]["baseScore"]
            else:
                cvss2 = 0
            if "weaknesses" in cve_data:
                cwes = cve_data["weaknesses"][0]
            else:
                cwes = {}

            if "configurations" in cve_data:
                vendors_products = cve_data["configurations"][0]
            else:
                vendors_products = {}
            # In case of multiple languages, keep the EN one
            descriptions = cve_data["descriptions"]
            if len(descriptions) > 1:
                descriptions = [d for d in descriptions if d["lang"] in ("en", "en-US")]
            summary = descriptions[0]["value"]
            cve = dict(  # noqa: C408
                cve_id=cve_id,
                summary=summary,
                json=cve_data,
                vendors=vendors_products,
                cwes=cwes,
                cvss2=cvss2,
                cvss3=cvss3,
            )
            # Create the CVEs mappings
            mappings["cves"].append(
                cve,
            )

        start_index += 2000

        # Update the objects in database
        if (start_index % 20_000 == 0) or (start_index >= total_results):
            for cve in mappings["cves"]:
                nvd_in = Nvd(**cve)
                nvd_in_hash_sum = hashlib.sha256(
                    json.dumps(nvd_in.model_dump(), sort_keys=True).encode("utf-8"),
                ).hexdigest()
                # print(nvd_in_hash_sum)
                server_data = requests.get(settings.CVE_API_URL + f"api/v1/cve/?cve_id={cve['cve_id']}").json()
                print(server_data)
                if not server_data["nvd"]:
                    resp = requests.post(settings.CVE_API_URL + "api/v1/nvd/", data=json.dumps(cve))
                    continue
                if server_data["nvd"]["hash_sum"] != nvd_in_hash_sum:
                    print("New data")
                    resp = requests.put(
                        settings.CVE_API_URL + f'api/v1/nvd/{server_data["nvd"]["id"]}',
                        data=json.dumps(cve),
                    )
                    print(resp.text)
                else:
                    print("Old data")
            mappings["cves"] = []
            mappings["changes"] = []

        if start_index <= total_results:
            # info("Waiting 6 seconds")
            time.sleep(6)


def update_files(root: Tree, level: int = 0) -> None:
    print("Start updating")
    session = requests.Session()
    for entry in root:
        if entry.type == "blob" and entry.name.lower().startswith("cve"):
            print(entry.name)
            # print("cve" in entry.name.lower())
            pocs, refs = extract_links_from_file(entry.data_stream)
            tmp = {
                "cve_id": entry.name.split(".")[0],
                "pocs": pocs,
                "references": refs,
            }
            poc = Cve(**tmp)
            cve_in_hash_sum = hashlib.sha256(
                json.dumps(poc.model_dump(), sort_keys=True).encode("utf-8"),
            ).hexdigest()
            server_data = session.get(settings.CVE_API_URL + f"api/v1/poc/cve_id/{tmp['cve_id']}")
            if server_data.status_code == http.HTTPStatus.NOT_FOUND:
                print(poc.model_dump_json())
                resp: requests.Response = session.post(
                    settings.CVE_API_URL + "api/v1/poc/", data=poc.model_dump_json().encode("UTF-8"),
                )
                print(f"{resp.text=} {resp.status_code=}" )
                continue
            if server_data.json()["hash_sum"] != cve_in_hash_sum:
                print("New data")
                resp = session.put(
                    settings.CVE_API_URL + f'api/v1/cwe/{server_data["id"]}',
                    data=json.dumps(tmp),
                )
                print(resp.text)
            else:
                print("Old data")
            # resp = requests.post(settings.CVE_API_URL + "api/v1/poc/", data=json.dumps(tmp))
            # print(pocs, "\n", refs)
        elif entry.type == "tree":
            update_files(entry, level + 1)


@celery.task
def update_poc() -> None:
    directory_to_repo = "./tmp"
    repo = Repo.clone_from(settings.REPO_URL, directory_to_repo)
    # prev_commits = list(repo.iter_commits(all=True, max_count=10))  # Last 10 commits from all branches.
    # tree = prev_commits[0].tree
    tree: Tree = repo.head.commit.tree
    update_files(tree)
    rmtree(directory_to_repo)


@celery.task
def update() -> None:
    update_poc()
    update_nvd()
    update_bdu()
    update_cwe()
