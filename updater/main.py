import hashlib
import json
import re
import time
from io import BytesIO
from zipfile import ZipFile

import requests
import untangle
from openpyxl import load_workbook

from config import settings

# from models import Nvd

def init_cwe():
    """
    Import the CWE list.
    """
    print("Importing CWE list...")

    # Download the file
    # with timed_operation("Downloading {}...".format(settings.MITRE_CWE_URL)):
    resp = requests.get(settings.MITRE_CWE_URL).content

    # Parse weaknesses
    # with timed_operation("Parsing cwes..."):
    z = ZipFile(BytesIO(resp))
    raw = z.open(z.namelist()[0]).read()
    obj = untangle.parse(raw.decode("utf-8"))
    weaknesses = obj.Weakness_Catalog.Weaknesses.Weakness
    categories = obj.Weakness_Catalog.Categories.Category

    # Create the objects
    cwes = {}
    # with timed_operation("Creating mappings..."):
    count = 0
    for c in weaknesses + categories:
        tmp = {
            # "id": get_uuid(),
            "cwe_id": f"CWE-{c['ID']}",
            "name": c["Name"],
            "description": c.Description.cdata
            if hasattr(c, "Description")
            else c.Summary.cdata,
        }
        resp = requests.post(settings.CVE_API_URL + "api/v1/cwe/", data=json.dumps(tmp))
        count += 1
        # print(cwes[c["ID"]])

    # Insert the objects in database
    # with timed_operation("Inserting CWE..."):
        # db.session.bulk_insert_mappings(Cwe, cwes.values())
        # db.session.commit()
    # for cwe in cwes:
    #     print(cwe)
    print(f"{count} CWE imported.")
    del cwes


def init_bdu() -> None:
    print("Getting info from BDU")
    # r = requests.get(settings.CVE_API_URL)
    # print(r)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    }
    xlsx = requests.get(settings.BDU_XLSX_URL, verify=False, allow_redirects=True, headers=headers).content  # noqa: S501
    wb = load_workbook(filename=BytesIO(xlsx))
    results = []
    sheet = wb.active
    session = requests.Session()
    # f = open('res.txt', 'w')
    for row in sheet.iter_rows(min_row=4):
        bdu_id = row[0].value
        description = row[1].value
        other_id = row[18].value
        # Создаем новый словарь для каждой итерации
        bdu_data = {"cve_id": "", "bdu_id": bdu_id, "description": description}
        if other_id:
            cve_match = re.findall(r"CVE-\d{4}-\d{4,7}", str(other_id))
            for cve_id in cve_match:
                # Создаем новый словарь для каждого cve_id
                temp_bdu_data = bdu_data.copy()
                temp_bdu_data["cve_id"] = cve_id
                results.append(temp_bdu_data)
        else:
            results.append(bdu_data)
    for bdu in results:
        _resp = session.post(settings.CVE_API_URL + "api/v1/bdu/", data=json.dumps(bdu))
        # print(resp)

def init_nvd() -> None:
    """Import the CVE list.

    Important notice:
        This product uses data from the NVD API but is not endorsed or certified by the NVD.
    """
    print("Gettins info from NVD")
    mappings = {"vendors": {}, "products": {}, "cves": [], "changes": []}
    url_template = settings.NVD_API_URL + "?startIndex={idx}"

    start_index = 0
    total_results = 0

    while start_index <= total_results:
        url = url_template.format(idx=start_index)
        # with timed_operation(f"Downloading {url}"):
        print(f"Downloading {url}")
        resp = requests.get(url)

        if not resp.ok:
            print(f"Bad response: {resp.status_code}, sleeping before retrying")
            time.sleep(10)
            continue
        data = resp.json()
        total_results = data.get("totalResults")
        for vulnerability in data.get("vulnerabilities"):
            cve_data = vulnerability.get("cve")
            cve_id = cve_data["id"]
            if "cvssMetricV31" in cve_data["metrics"]:
                cvss3 = cve_data.get("metrics")["cvssMetricV31"][0]["cvssData"]["baseScore"]
            elif "cvssMetricV30" in cve_data["metrics"]:
                cvss3 = cve_data.get("metrics")["cvssMetricV30"][0]["cvssData"]["baseScore"]
            else:
                cvss3 = 0

            if "cvssMetricV2" in cve_data.get("metrics"):
                cvss2 = cve_data.get("metrics")["cvssMetricV2"][0]["cvssData"]["baseScore"]
            else:
                cvss2 = 0
            if "weaknesses" in cve_data:
                cwes = cve_data["weaknesses"][0]
            else:
                cwes = {}

            if "configurations" in cve_data:
                vendors_products = cve_data["configurations"][0]
            else:
                vendors_products = {}
            # In case of multiple languages, keep the EN one
            descriptions = cve_data["descriptions"]
            if len(descriptions) > 1:
                descriptions = [d for d in descriptions if d["lang"] in ("en", "en-US")]
            summary = descriptions[0]["value"]
            cve = dict(  # noqa: C408
                # id=cve_db_id,
                cve_id=cve_id,
                summary=summary,
                json=cve_data,
                vendors=vendors_products,
                cwes=cwes,
                cvss2=cvss2,
                cvss3=cvss3,
                # created_at=arrow.get(cve_data["published"]).datetime,
                # updated_at=arrow.get(cve_data["lastModified"]).datetime,
            )
            # Create the CVEs mappings
            mappings["cves"].append(
                cve,
            )
            # nvd_in = NvdCreat(**cve)
            # nvd_in_hash_sum = hashlib.sha256(json.dumps(nvd_in.model_dump(), sort_keys=True).encode("utf-8")).hexdigest()
            # print(nvd_in_hash_sum)
            # print(cve)
            # break

        # NVD requirement is 2000 CVE per page
        start_index += 2000

        # Insert the objects in database
        if (start_index % 20_000 == 0) or (start_index >= total_results):
            # with timed_operation("Inserting CVE"):
            # db.session.bulk_insert_mappings(Cve, mappings["cves"])
            # db.session.commit()
            # Create the changes based on CVEs data
            for cve in mappings["cves"]:
                # print(cve)
                resp = requests.post(settings.CVE_API_URL + "api/v1/nvd/", data=json.dumps(cve))
                # print(resp.text)
                # time.sleep(1)
                # mappings["changes"].append(
                #     dict(
                #         id=get_uuid(),
                #         created_at=cve["created_at"],
                #         updated_at=cve["updated_at"],
                #         json=cve["json"],
                #         cve_id=cve["id"],
                #         task_id=task_id,
                #     )
                # )
            # db.session.bulk_insert_mappings(Change, mappings["changes"])
            # db.session.commit()

            # info("{} CVE imported.".format(len(mappings["cves"])))

            # Free the memory after each processed year
            mappings["cves"] = []
            mappings["changes"] = []

        # NVD requirement is 6s between requests
        if start_index <= total_results:
            # info("Waiting 6 seconds")
            time.sleep(6)



if __name__ == "__main__":
    # init_bdu()
    # init_nvd()
    init_cwe()
